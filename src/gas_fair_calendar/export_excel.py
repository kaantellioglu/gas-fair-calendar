from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from .io_json import load_events
from .settings import DATA_DIR, OUTPUT_DIR


HEADERS = [
    "ID",
    "Name",
    "Short Name",
    "Start Date",
    "End Date",
    "City",
    "Country",
    "Region",
    "Category",
    "Venue",
    "Organizer",
    "Website",
    "Ticket Info",
    "Status",
    "Confidence Score",
    "Source URL",
    "Source Type",
    "Last Checked At",
]


def export_excel() -> Path:
    events = sorted(load_events(DATA_DIR / "events_master.json"), key=lambda e: (e.start_date, e.name))
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    ws.append(HEADERS)

    header_fill = PatternFill(fill_type="solid", start_color="1F4E78", end_color="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for event in events:
        ws.append([
            event.id,
            event.name,
            event.short_name,
            event.start_date.isoformat(),
            event.end_date.isoformat(),
            event.city,
            event.country,
            event.region,
            event.category,
            event.venue,
            event.organizer,
            event.website,
            event.ticket_info,
            event.status,
            event.confidence_score,
            event.source_url,
            event.source_type,
            event.last_checked_at.isoformat() if event.last_checked_at else "",
        ])

    widths = {
        "A": 18, "B": 36, "C": 22, "D": 14, "E": 14, "F": 18, "G": 22,
        "H": 12, "I": 12, "J": 28, "K": 24, "L": 38, "M": 22, "N": 16,
        "O": 16, "P": 40, "Q": 18, "R": 24,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    out = OUTPUT_DIR / "events_master.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    path = export_excel()
    print(f"Excel created: {path}")
